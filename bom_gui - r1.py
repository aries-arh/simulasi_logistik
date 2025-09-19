import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, ttk
from bom_parser import parse_bom_file, print_bom_hierarchy, get_bom_hierarchy_list, load_material_data
import openpyxl
import json # Tambahkan import json
import pyperclip # Untuk copy ke clipboard
import os
import shutil
import glob

# Kelas untuk tampilan BOM Viewer yang sudah ada
class BOMViewerApp:
    def __init__(self, master):
        self.master = master
        self.master.title("BOM Hierarchy Viewer")
        # Atur ukuran jendela dan posisikan di tengah layar
        window_width = 800
        window_height = 600
        screen_width = master.winfo_screenwidth()
        screen_height = master.winfo_screenheight()
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        master.geometry(f'{window_width}x{window_height}+{x}+{y}')

        self.material_data = {} # Tambahkan untuk menyimpan data material
        self.load_material_plant_data()

        self.bom_data = {}
        # self.load_material_plant_data() # Tidak perlu lagi, material_data diterima

        # Frame untuk input file
        self.file_frame = tk.LabelFrame(master, text="File BOM")
        self.file_frame.pack(padx=10, pady=5, fill="x")

        self.file_path_label = tk.Label(self.file_frame, text="Jalur File:")
        self.file_path_label.pack(side="left", padx=5, pady=5)

        self.file_path_entry = tk.Entry(self.file_frame, width=50)
        self.file_path_entry.pack(side="left", padx=5, pady=5, expand=True, fill="x")

        self.browse_button = tk.Button(self.file_frame, text="Browse", command=self.browse_file)
        self.browse_button.pack(side="left", padx=5, pady=5)

        # Frame untuk input part induk
        self.part_frame = tk.LabelFrame(master, text="Part Induk")
        self.part_frame.pack(padx=10, pady=5, fill="x")

        self.part_label = tk.Label(self.part_frame, text="Masukkan Part Induk:")
        self.part_label.pack(side="left", padx=5, pady=5)

        self.part_entry = tk.Entry(self.part_frame, width=30)
        self.part_entry.pack(side="left", padx=5, pady=5, expand=True, fill="x")

        self.view_button = tk.Button(self.part_frame, text="Tampilkan Hirarki", command=self.display_hierarchy)
        self.view_button.pack(side="left", padx=5, pady=5)

        self.export_button = tk.Button(self.part_frame, text="Export to Excel", command=self.export_to_excel)
        self.export_button.pack(side="left", padx=5, pady=5)

        # Area tampilan hirarki
        self.result_frame = tk.LabelFrame(master, text="Hasil Hirarki")
        self.result_frame.pack(padx=10, pady=5, fill="both", expand=True)

        self.hierarchy_text = scrolledtext.ScrolledText(self.result_frame, wrap=tk.WORD, width=80, height=20)
        self.hierarchy_text.pack(padx=5, pady=5, fill="both", expand=True)

    def load_material_plant_data(self):
        # Path ke file material plant data list
        material_file_path = "material_plant_data.txt"
        try:
            self.material_data = load_material_data(material_file_path)
            if self.material_data:
                print(f"Data Material Plant berhasil dimuat. Jumlah entri: {len(self.material_data)}")
            else:
                print("Gagal memuat data Material Plant atau file kosong.")
        except Exception as e:
            print(f"Terjadi kesalahan saat memuat data material plant: {e}")

    def browse_file(self):
        file_selected = filedialog.askopenfilename(
            title="Pilih File BOM",
            filetypes=(("Text files", "*.txt"), ("All files", "*.*"))
        )
        if file_selected:
            self.file_path_entry.delete(0, tk.END)
            self.file_path_entry.insert(0, file_selected)
            self.load_bom_data(file_selected)

    def load_bom_data(self, file_path):
        self.hierarchy_text.delete(1.0, tk.END)
        self.hierarchy_text.insert(tk.END, f"Memuat data dari {file_path}...\n")
        self.bom_data = parse_bom_file(file_path)
        if self.bom_data:
            self.hierarchy_text.insert(tk.END, "Data BOM berhasil dimuat.\n")
        else:
            self.hierarchy_text.insert(tk.END, "Gagal memuat data BOM atau file kosong.\n")

    def display_hierarchy(self):
        self.hierarchy_text.delete(1.0, tk.END)
        part_induk = self.part_entry.get().strip()

        if not self.bom_data:
            self.hierarchy_text.insert(tk.END, "Silakan muat file BOM terlebih dahulu.\n")
            
            return

        if not part_induk:
            self.hierarchy_text.insert(tk.END, "Silakan masukkan Part Induk.\n")
            return
        
        # Cek apakah part induk ada di bom_data sebagai parent atau sebagai child dari parent lain
        if part_induk in self.bom_data or any(part_induk in children for children in self.bom_data.values()):
            self.hierarchy_text.insert(tk.END, f"Part Induk: {part_induk}\n")
            self.hierarchy_text.insert(tk.END, "\nHirarki Anak:\n")

            # Jika part induk memiliki anak langsung
            if part_induk in self.bom_data:
                for child in self.bom_data[part_induk]:
                    print_bom_hierarchy(self.bom_data, child, indent=0, text_widget=self.hierarchy_text, material_data=self.material_data)
            else:
                # Menangani kasus di mana part_induk tidak memiliki anak langsung
                # tetapi mungkin ada di BOM sebagai anak dari part lain
                self.hierarchy_text.insert(tk.END, f"Part \'{part_induk}\' tidak memiliki anak langsung dalam data BOM.\n")

        else:
            self.hierarchy_text.insert(tk.END, f"Part Induk \'{part_induk}\' tidak ditemukan dalam data BOM.\n")

    def export_to_excel(self):
        if not self.bom_data:
            messagebox.showwarning("Peringatan", "Tidak ada data BOM untuk diekspor. Harap muat file BOM terlebih dahulu.")
            return

        part_induk = self.part_entry.get().strip()
        if not part_induk:
            messagebox.showwarning("Peringatan", "Silakan masukkan Part Induk untuk diekspor hirarkinya.")
            return

        if part_induk not in self.bom_data and not any(part_induk in children for children in self.bom_data.values()):
            messagebox.showwarning("Peringatan", f"Part Induk \'{part_induk}\' tidak ditemukan dalam data BOM.")
            return

        # Kumpulkan data hirarki
        hierarchy_data = get_bom_hierarchy_list(self.bom_data, part_induk, self.material_data)

        if not hierarchy_data:
            messagebox.showinfo("Informasi", f"Tidak ada hirarki yang ditemukan untuk Part Induk \'{part_induk}'.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
            title="Simpan Hirarki BOM sebagai Excel"
        )

        if not file_path:
            return # Pengguna membatalkan

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "BOM Hierarchy"

            # Tulis Part Induk yang diinputkan secara terpisah
            sheet['A1'] = "Part Induk:"
            sheet['B1'] = part_induk

            # Tulis header untuk hirarki anak
            sheet['A2'] = "Level No."
            sheet['B2'] = "Component"
            sheet['C2'] = "Material Description"
            sheet['D2'] = "MRP Type"
            sheet['E2'] = "MRPC"

            # Kumpulkan data hirarki anak (mulai dari level 1 untuk anak langsung)
            children_hierarchy_data = []
            if part_induk in self.bom_data:
                for child in self.bom_data[part_induk]:
                    # Panggil get_bom_hierarchy_list untuk setiap anak langsung
                    # dengan indent awal 0 agar mereka mulai dari .1
                    children_hierarchy_data.extend(get_bom_hierarchy_list(self.bom_data, child, indent=0, material_data=self.material_data))
            
            # Tulis data hirarki anak dengan data tambahan
            for i, (level, part, material_info) in enumerate(children_hierarchy_data):
                sheet.cell(row=i + 3, column=1, value=level)
                sheet.cell(row=i + 3, column=2, value=part)

                # Lakukan VLOOKUP pada self.material_data
                # material_info = self.material_data.get(part, {}) # Tidak perlu lagi, sudah ada di hierarchy_data
                # print(f"Mencari Part: {part}, Info Material: {material_info}") # Debugging print, bisa dihapus jika tidak diperlukan
                sheet.cell(row=i + 3, column=3, value=material_info.get('Material Description', ''))
                sheet.cell(row=i + 3, column=4, value=material_info.get('MRP Type', ''))
                sheet.cell(row=i + 3, column=5, value=material_info.get('MRPC', ''))
            
            workbook.save(file_path)
            messagebox.showinfo("Sukses", f"Hirarki BOM berhasil diekspor ke:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Terjadi kesalahan saat mengekspor ke Excel:\n{e}")

# Kelas utama aplikasi
class MainMenuApp:
    def __init__(self, master):
        self.master = master
        master.title("Aplikasi BOM & Master Data")


        # Atur ukuran jendela dan posisikan di tengah layar
        window_width = 800
        window_height = 600
        screen_width = master.winfo_screenwidth()
        screen_height = master.winfo_screenheight()
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        master.geometry(f'{window_width}x{window_height}+{x}+{y}')

        self.material_data = {} # Simpan data material di sini
        self.load_material_plant_data()

        # Instance Data Line Manager tanpa GUI untuk mendukung Data Routing dari menu utama
        self.data_line_manager_instance = DataLineManager(None, init_gui=False)

        # Frame untuk tombol-tombol menu utama
        self.main_frame = tk.Frame(master)
        self.main_frame.pack(padx=20, pady=20)

        self.master_data_button = tk.Button(self.main_frame, text="Master Data", command=self.open_master_data_window)
        self.master_data_button.pack(pady=10)

        # Tambahkan tombol Data Routing di menu utama
        self.data_routing_button_main = tk.Button(self.main_frame, text="Data Routing", command=self.open_data_routing_from_main)
        self.data_routing_button_main.pack(pady=10)

        # Contoh: Jika ingin tetap ada tombol Cek BOM di main menu juga
        # self.cek_bom_button_main = tk.Button(self.main_frame, text="Cek BOM (Langsung)", command=self.open_bom_viewer_direct)
        # self.cek_bom_button_main.pack(pady=10)

    def load_material_plant_data(self, show_messages=False):
        sumber_folder = "N:\\Download\\MaterialPlantDataList"
        tujuan_file = "material_plant_data.txt"
        file_copied = False

        try:
            # 1. Cari file terbaru di folder sumber
            list_of_files = glob.glob(os.path.join(sumber_folder, '*.txt'))
            if list_of_files:
                file_terbaru = max(list_of_files, key=os.path.getmtime)
                # 2. Salin file terbaru ke tujuan
                shutil.copy(file_terbaru, tujuan_file)
                file_copied = True
                print(f"File material terbaru '{os.path.basename(file_terbaru)}' berhasil disalin.")
            elif show_messages:
                messagebox.showwarning("Peringatan", f"Tidak ada file .txt ditemukan di {sumber_folder}.")

        except Exception as e:
            print(f"Gagal menyalin file dari jaringan: {e}")
            if show_messages:
                messagebox.showerror("Error Jaringan", f"Tidak dapat mengakses folder sumber: {sumber_folder}.\n\nError: {e}\n\nAplikasi akan mencoba menggunakan data lokal yang ada.")

        # 3. Muat data dari file lokal
        try:
            self.material_data = load_material_data(tujuan_file)
            if self.material_data:
                print(f"Data Material Plant berhasil dimuat. Jumlah entri: {len(self.material_data)}")
                if show_messages and file_copied:
                    messagebox.showinfo("Sukses", "Data material terbaru berhasil dimuat.")
            else:
                print("Gagal memuat data Material Plant atau file lokal kosong.")
                if show_messages:
                    messagebox.showwarning("Peringatan", "Gagal memuat data dari file lokal atau file kosong.")
        except FileNotFoundError:
            print(f"File lokal '{tujuan_file}' tidak ditemukan.")
            self.material_data = {} # Pastikan material_data adalah dict kosong
            if show_messages:
                messagebox.showerror("Error", f"File data material lokal '{tujuan_file}' tidak ditemukan. Harap muat data terlebih dahulu.")
        except Exception as e:
            print(f"Terjadi kesalahan saat memuat data material plant: {e}")
            self.material_data = {} # Pastikan material_data adalah dict kosong
            if show_messages:
                messagebox.showerror("Error", f"Terjadi kesalahan saat memuat data material: {e}")

    def reload_material_data(self):
        # Fungsi untuk memuat ulang data material dengan message box
        self.load_material_plant_data(show_messages=True)

    def open_master_data_window(self):
        master_data_root = tk.Toplevel(self.master)
        master_data_root.transient(self.master) # Buat jendela master data transient
        master_data_root.grab_set() # Buat jendela master data modal
        master_data_app = MasterDataWindow(master_data_root, self.material_data, self) # Lewatkan MainMenuApp instance
        self.master.wait_window(master_data_root) # Tunggu sampai jendela master data ditutup

    def open_bom_viewer_direct(self):
        bom_viewer_root = tk.Toplevel(self.master)
        BOMViewerApp(bom_viewer_root) # Tidak lagi meneruskan material_data

    def open_data_routing_from_main(self):
        data_routing_root = tk.Toplevel(self.master)
        data_routing_root.transient(self.master)
        data_routing_root.grab_set()
        DataRoutingManager(data_routing_root, self.material_data, self.data_line_manager_instance, self)
        self.master.wait_window(data_routing_root)

# Kelas untuk jendela Master Data
class MasterDataWindow:
    def __init__(self, master, material_data, main_app_instance):
        self.master = master
        self.master.title("Jendela Master Data")

        # Atur ukuran jendela dan posisikan di tengah layar
        window_width = 700 # Ukuran lebar yang diinginkan
        window_height = 500 # Ukuran tinggi yang diinginkan
        screen_width = master.winfo_screenwidth()
        screen_height = master.winfo_screenheight()
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        master.geometry(f'{window_width}x{window_height}+{x}+{y}')

        self.material_data = material_data # Simpan material_data
        self.main_app_instance = main_app_instance # Simpan referensi ke MainMenuApp
        self.data_line_manager_instance = DataLineManager(None, init_gui=False) # Inisialisasi tanpa GUI

        self.master_frame = tk.Frame(master)
        self.master_frame.pack(padx=20, pady=20)

        # Hapus tombol Cek BOM dan Data Line dari jendela Master Data
        # Data Routing dipindahkan ke tampilan awal (Main Menu)

        # self.muat_material_button = tk.Button(self.master_frame, text="Muat Material Plant Data List", command=self.muat_material_plant_data)
        # self.muat_material_button.pack(pady=10)

        # self.update_mrp_button = tk.Button(self.master_frame, text="Update MRP-Controller Data", command=self.update_mrp_controller_data) # Tombol baru
        # self.update_mrp_button.pack(pady=10)

    # def update_mrp_controller_data(self):
    #     self.main_app_instance.reload_material_data()
    #     self.material_data = self.main_app_instance.material_data # Perbarui material_data di MasterDataWindow

    def open_bom_viewer(self):
        bom_viewer_root = tk.Toplevel(self.master)
        bom_viewer_root.transient(self.master)
        bom_viewer_root.grab_set()
        BOMViewerApp(bom_viewer_root) # Tidak lagi meneruskan material_data
        self.master.wait_window(bom_viewer_root)

    def open_data_line_manager(self):
        data_line_root = tk.Toplevel(self.master)
        data_line_root.transient(self.master)
        data_line_root.grab_set()
        # Buat instance DataLineManager baru, tetapi lewati data_lines yang sudah ada
        DataLineManager(data_line_root, existing_data_lines=self.data_line_manager_instance.data_lines, material_data=self.material_data, main_app_instance=self.main_app_instance) # Meneruskan material_data dan main_app_instance
        self.master.wait_window(data_line_root)

    def open_data_routing_manager(self):
        data_routing_root = tk.Toplevel(self.master)
        data_routing_root.transient(self.master)
        data_routing_root.grab_set()
        DataRoutingManager(data_routing_root, self.material_data, self.data_line_manager_instance, self.main_app_instance) # Mengembalikan panggilan ini
        self.master.wait_window(data_routing_root)

    # def muat_material_plant_data(self):
    #     if self.main_app_instance:
    #         self.main_app_instance.reload_material_data()
    #         # Perbarui referensi material_data di jendela ini juga
    #         self.material_data = self.main_app_instance.material_data
    #         # Anda mungkin perlu memperbarui tampilan lain yang bergantung pada data ini
    #         # contohnya, jika ada listbox yang menampilkan data dari material_data
    #         # panggil method untuk me-refresh listbox tersebut di sini.
    #         # Misalnya, jika DataRoutingManager perlu di-refresh:
    #         # (Ini memerlukan cara untuk mengakses instance DataRoutingManager yang aktif)
    #         messagebox.showinfo("Info", "Proses pembaruan data material selesai.")
    #     else:
    #         messagebox.showerror("Error", "Referensi ke aplikasi utama tidak ditemukan.")

# Kelas untuk mengelola Data Line
class DataLineManager:
    def __init__(self, master, existing_data_lines=None, init_gui=True, material_data=None, main_app_instance=None):
        self.master = master
        self.data_lines = existing_data_lines if existing_data_lines is not None else []
        self.material_data = material_data if material_data is not None else {}
        self.main_app_instance = main_app_instance
        
        if init_gui:
            self.master.title("Manajemen Data Line")
            # Atur ukuran jendela dan posisikan di tengah layar
            window_width = 700
            window_height = 500
            screen_width = master.winfo_screenwidth()
            screen_height = master.winfo_screenheight()
            x = (screen_width // 2) - (window_width // 2)
            y = (screen_height // 2) - (window_height // 2)
            master.geometry(f'{window_width}x{window_height}+{x}+{y}')

            self.load_data_lines() # Muat data line yang sudah ada

            self.main_frame = tk.Frame(master)
            self.main_frame.pack(padx=20, pady=20)

            # Input untuk MRP Controller
            tk.Label(self.main_frame, text="MRP Controller:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
            self.mrp_controller_line_entry = tk.Entry(self.main_frame, width=30)
            self.mrp_controller_line_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
            self.mrp_controller_line_entry.bind("<KeyRelease>", self.autofill_mrp_controller_line)

            self.mrp_controller_line_listbox = tk.Listbox(self.main_frame, height=5)
            self.mrp_controller_line_listbox.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
            self.mrp_controller_line_listbox.bind('<<ListboxSelect>>', self.select_mrp_controller_line)
            self.mrp_controller_line_listbox.grid_remove() # Sembunyikan secara default

            # Input untuk Nama Line
            tk.Label(self.main_frame, text="Nama Line:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
            self.line_name_entry = tk.Entry(self.main_frame, width=30)
            self.line_name_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

            # Tombol Tambah Line
            self.add_button = tk.Button(self.main_frame, text="Tambah Line", command=self.add_line)
            self.add_button.grid(row=2, column=2, padx=5, pady=5)

            # Tombol Update MRP-Controller Data
            self.update_mrp_button = tk.Button(self.main_frame, text="Update MRP-Controller Data", command=self.update_mrp_controller_data)
            self.update_mrp_button.grid(row=0, column=4, padx=5, pady=5) # Mengubah kolom dari 3 ke 4

            # Tombol Tampilkan Semua Line atau Filter Line
            self.show_lines_button = tk.Button(self.main_frame, text="Tampilkan/Filter Lines", command=self.refresh_line_list) # Menggunakan kembali refresh_line_list
            self.show_lines_button.grid(row=2, column=3, padx=5, pady=5)

            # Tampilan Data Line
            self.line_list_frame = tk.LabelFrame(self.main_frame, text="Daftar Data Line")
            self.line_list_frame.grid(row=3, column=0, columnspan=3, padx=5, pady=10, sticky="nsew")

            self.line_listbox = tk.Listbox(self.line_list_frame, height=10)
            self.line_listbox.pack(side="left", fill="both", expand=True)

            self.scrollbar = tk.Scrollbar(self.line_list_frame, orient="vertical", command=self.line_listbox.yview)
            self.scrollbar.pack(side="right", fill="y")
            self.line_listbox.config(yscrollcommand=self.scrollbar.set)

            # Tombol Hapus Line
            self.delete_button = tk.Button(self.main_frame, text="Hapus Line Terpilih", command=self.delete_selected_line)
            self.delete_button.grid(row=4, column=0, columnspan=3, pady=10)

            # Tampilan Daftar MRP Controller
            self.mrp_controller_display_frame = tk.LabelFrame(self.main_frame, text="Daftar MRP Controller")
            self.mrp_controller_display_frame.grid(row=1, column=4, rowspan=4, padx=5, pady=10, sticky="nsew") # Menyesuaikan rowspan

            self.mrp_controller_list_box = tk.Listbox(self.mrp_controller_display_frame, height=15)
            self.mrp_controller_list_box.pack(side="left", fill="both", expand=True)

            self.mrp_controller_scrollbar = tk.Scrollbar(self.mrp_controller_display_frame, orient="vertical", command=self.mrp_controller_list_box.yview)
            self.mrp_controller_scrollbar.pack(side="right", fill="y")
            self.mrp_controller_list_box.config(yscrollcommand=self.mrp_controller_scrollbar.set)

            self.refresh_line_list()
            self.refresh_mrp_controller_list()
    
    def update_mrp_controller_data(self):
        if self.main_app_instance:
            self.main_app_instance.reload_material_data()
            self.material_data = self.main_app_instance.material_data # Perbarui material_data di DataLineManager
            self.refresh_mrp_controller_list()
        else:
            messagebox.showwarning("Peringatan", "Aplikasi utama tidak tersedia untuk memperbarui data material.")

    def refresh_mrp_controller_list(self):
        self.mrp_controller_list_box.delete(0, tk.END)
        unique_mrpc = set()
        for material_info in self.material_data.values():
            mrpc_val = material_info.get('MRPC', '')
            if mrpc_val:
                unique_mrpc.add(mrpc_val)
        
        for mrpc in sorted(list(unique_mrpc)):
            self.mrp_controller_list_box.insert(tk.END, mrpc)

    def load_data_lines(self):
        # Memuat data line dari file (jika ada)
        try:
            with open("data_lines.json", "r") as f:
                self.data_lines = json.load(f)
        except FileNotFoundError:
            self.data_lines = []
        except json.JSONDecodeError:
            messagebox.showwarning("Peringatan", "File data_lines.json rusak atau kosong. Membuat file baru.")
            self.data_lines = []

    def save_data_lines(self):
        # Menyimpan data line ke file
        with open("data_lines.json", "w") as f:
            json.dump(self.data_lines, f, indent=4)

    def get_next_line_number(self):
        if not self.data_lines:
            return 1
        return max(line['number'] for line in self.data_lines) + 1

    def add_line(self):
        line_name = self.line_name_entry.get().strip()
        mrp_controller = self.mrp_controller_line_entry.get().strip()

        if not mrp_controller:
            messagebox.showwarning("Peringatan", "MRP Controller tidak boleh kosong.")
            return

        if line_name:
            next_num = self.get_next_line_number()
            self.data_lines.append({"number": next_num, "name": line_name, "mrp_controller": mrp_controller})
            self.save_data_lines()
            self.refresh_line_list() # Panggil tanpa argumen untuk menampilkan semua atau yang difilter
            self.line_name_entry.delete(0, tk.END)
            # self.mrp_controller_line_entry.delete(0, tk.END) # Biarkan MRP Controller tetap terpilih
        else:
            messagebox.showwarning("Peringatan", "Nama Line tidak boleh kosong.")

    def delete_selected_line(self):
        selected_indices = self.line_listbox.curselection()
        if selected_indices:
            index_to_delete = selected_indices[0]
            del self.data_lines[index_to_delete]
            self.save_data_lines()
            self.refresh_line_list() # Panggil tanpa argumen untuk menampilkan semua atau yang difilter
        else:
            messagebox.showwarning("Peringatan", "Pilih baris yang ingin dihapus.")

    def refresh_line_list(self):
        self.line_listbox.delete(0, tk.END)
        filter_mrpc = self.mrp_controller_line_entry.get().strip()

        lines_to_display = []
        if filter_mrpc:
            for line_data in self.data_lines:
                if line_data.get('mrp_controller') == filter_mrpc:
                    lines_to_display.append(line_data)
        else:
            lines_to_display = self.data_lines

        for line_data in lines_to_display:
            self.line_listbox.insert(tk.END, f"{line_data['number']}. {line_data['name']} (MRP: {line_data['mrp_controller']})")

    def autofill_mrp_controller_line(self, event=None):
        search_term = self.mrp_controller_line_entry.get().strip().upper()
        matching_mrpc = []
        for material_info in self.material_data.values():
            mrpc_val = material_info.get('MRPC', '')
            if search_term and mrpc_val.startswith(search_term) and mrpc_val not in matching_mrpc:
                matching_mrpc.append(mrpc_val)
        
        self.mrp_controller_line_listbox.delete(0, tk.END)
        if matching_mrpc:
            for mrpc in sorted(matching_mrpc):
                self.mrp_controller_line_listbox.insert(tk.END, mrpc)
            self.mrp_controller_line_listbox.grid()
        else:
            self.mrp_controller_line_listbox.grid_remove()

    def select_mrp_controller_line(self, event):
        if self.mrp_controller_line_listbox.curselection():
            selected_mrp = self.mrp_controller_line_listbox.get(self.mrp_controller_line_listbox.curselection())
            self.mrp_controller_line_entry.delete(0, tk.END)
            self.mrp_controller_line_entry.insert(0, selected_mrp)
            self.mrp_controller_line_listbox.grid_remove()
            self.refresh_line_list() # Memperbarui daftar line saat MRP Controller dipilih

# Kelas untuk mengelola Data Routing
class DataRoutingManager:
    def __init__(self, master, material_data, data_lines_manager, main_app_instance):
        self.master = master
        self.master.title("Manajemen Data Routing")
        self.material_data = material_data
        self.data_lines_manager = data_lines_manager # Untuk mengakses data lines
        self.main_app_instance = main_app_instance # Untuk mengakses instance MainMenuApp
        
        # Inisialisasi variabel BOM di awal
        self.bom_file_path = None
        self.bom_data = {}
        self.routings = {} # Pindahkan inisialisasi self.routings ke sini

        # Atur ukuran jendela dan posisikan di tengah layar
        window_width = 800
        window_height = 700
        screen_width = master.winfo_screenwidth()
        screen_height = master.winfo_screenheight()
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        master.geometry(f'{window_width}x{window_height}+{x}+{y}')

        self.main_frame = tk.Frame(master)
        self.main_frame.pack(padx=20, pady=20, fill="both", expand=True)
        
        # Konfigurasi grid weights agar elemen dapat berkembang
        self.main_frame.grid_columnconfigure(1, weight=1)
        self.main_frame.grid_rowconfigure(10, weight=1) # Update row configure ke row 10

        # Bagian Input MRP Controller
        # Tombol Muat Material Plant Data List di Data Routing
        self.muat_material_button = tk.Button(self.main_frame, text="Muat Material Plant Data List", command=self.muat_material_plant_data)
        self.muat_material_button.grid(row=0, column=0, padx=5, pady=5, sticky="w") # Pindahkan ke row 0, column 0

        # Tombol Muat File BOM di samping Muat Material Plant Data List
        self.load_bom_file_button = tk.Button(self.main_frame, text="Muat File BOM", command=self.load_bom_file)
        self.load_bom_file_button.grid(row=0, column=1, padx=5, pady=5, sticky="w") # Pindahkan ke row 0, column 1

        tk.Label(self.main_frame, text="MRP Controller:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.mrp_controller_entry = tk.Entry(self.main_frame, width=30)
        self.mrp_controller_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        self.mrp_controller_entry.bind("<KeyRelease>", self.autofill_mrp_controller)

        self.mrp_controller_listbox = tk.Listbox(self.main_frame, height=5)
        self.mrp_controller_listbox.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
        self.mrp_controller_listbox.bind('<<ListboxSelect>>', self.select_mrp_controller)
        self.mrp_controller_listbox.grid_remove() # Sembunyikan secara default

        # Tombol untuk menampilkan parts terkait MRP Controller
        self.show_parts_button = tk.Button(self.main_frame, text="Tampilkan Parts", command=self.display_parts_for_mrp_controller)
        self.show_parts_button.grid(row=1, column=2, padx=5, pady=5) # Sesuaikan baris

        # Bagian Tampilan Parts (diubah menjadi Treeview)
        self.parts_list_frame = tk.LabelFrame(self.main_frame, text="Parts Terkait")
        self.parts_list_frame.grid(row=3, column=0, columnspan=3, padx=5, pady=10, sticky="nsew")

        self.parts_tree = ttk.Treeview(self.parts_list_frame, columns=("Part", "Description", "SPT"), show="headings", height=5)
        self.parts_tree.heading("Part", text="Part")
        self.parts_tree.heading("Description", text="Material Description")
        self.parts_tree.heading("SPT", text="SPT")
        self.parts_tree.column("Part", width=150)
        self.parts_tree.column("Description", width=250)
        self.parts_tree.column("SPT", width=100)

        self.parts_tree.pack(side="left", fill="both", expand=True)

        self.parts_scrollbar = ttk.Scrollbar(self.parts_list_frame, orient="vertical", command=self.parts_tree.yview)
        self.parts_scrollbar.pack(side="right", fill="y")
        self.parts_tree.config(yscrollcommand=self.parts_scrollbar.set)
        self.parts_tree.bind('<<TreeviewSelect>>', self.select_part_from_list)

        # Input untuk Part yang Dipilih dan MRP Type-nya
        tk.Label(self.main_frame, text="Part Dipilih:").grid(row=4, column=0, padx=5, pady=5, sticky="w") # Pindahkan ke row 4
        self.selected_part_entry = tk.Entry(self.main_frame, width=30)
        self.selected_part_entry.grid(row=4, column=1, padx=5, pady=5, sticky="ew") # Pindahkan ke row 4
        self.selected_part_entry.bind("<KeyRelease>", self.autofill_selected_part) # Tambahkan autofill untuk input part

        self.selected_part_listbox = tk.Listbox(self.main_frame, height=5) # Listbox untuk autofill part
        self.selected_part_listbox.grid(row=5, column=1, padx=5, pady=5, sticky="ew") # Pindahkan ke row 5
        self.selected_part_listbox.bind('<<ListboxSelect>>', self.select_selected_part)
        self.selected_part_listbox.grid_remove() # Sembunyikan secara default

        tk.Label(self.main_frame, text="MRP Type Part:").grid(row=6, column=0, padx=5, pady=5, sticky="w") # Pindahkan ke row 6
        self.selected_part_mrp_type_entry = tk.Entry(self.main_frame, width=30, state="readonly") # Hanya untuk tampilan
        self.selected_part_mrp_type_entry.grid(row=6, column=1, padx=5, pady=5, sticky="ew") # Pindahkan ke row 6

        # Tombol Muat dan Tampilkan BOM (dipindahkan ke sini)
        self.bom_buttons_frame_top = tk.Frame(self.main_frame)
        self.bom_buttons_frame_top.grid(row=7, column=0, columnspan=3, pady=10, sticky="w")

        self.load_bom_file_button_top = tk.Button(self.bom_buttons_frame_top, text="Muat File BOM", command=self.load_bom_file)
        self.load_bom_file_button_top.pack(side="left", padx=5)

        self.display_bom_button_top = tk.Button(self.bom_buttons_frame_top, text="Tampilkan BOM", command=lambda: self.display_bom_hierarchy_for_part(self.selected_part_entry.get().strip()))
        self.display_bom_button_top.pack(side="left", padx=5)

        # Frame baru untuk filter anak langsung
        self.first_level_filter_frame = tk.LabelFrame(self.main_frame, text="Filter Anak Langsung")
        self.first_level_filter_frame.grid(row=8, column=0, columnspan=3, padx=10, pady=5, sticky="nsew")

        # Konfigurasi grid untuk elemen di dalam first_level_filter_frame
        self.first_level_filter_frame.grid_columnconfigure(1, weight=1) # Agar entry filter bisa melebar
        self.first_level_filter_frame.grid_columnconfigure(3, weight=1) # Agar entry filter bisa melebar

        # Filter MRPC
        tk.Label(self.first_level_filter_frame, text="Filter MRPC:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.first_level_mrpc_filter_entry = tk.Entry(self.first_level_filter_frame, width=15)
        self.first_level_mrpc_filter_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.first_level_mrpc_filter_entry.bind("<KeyRelease>", self.autofill_first_level_mrpc)

        self.first_level_mrpc_listbox = tk.Listbox(self.first_level_filter_frame, height=3)
        self.first_level_mrpc_listbox.grid(row=1, column=1, padx=5, pady=0, sticky="ew")
        self.first_level_mrpc_listbox.bind('<<ListboxSelect>>', self.select_first_level_mrpc)
        self.first_level_mrpc_listbox.grid_remove() # Sembunyikan secara default

        # Filter SPT
        tk.Label(self.first_level_filter_frame, text="Filter SPT:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.first_level_spt_filter_entry = tk.Entry(self.first_level_filter_frame, width=15)
        self.first_level_spt_filter_entry.grid(row=0, column=3, padx=5, pady=5, sticky="ew")
        self.first_level_spt_filter_entry.insert(0, "*")  # Set default value to "*" (show all SPT)
        self.first_level_spt_filter_entry.bind("<KeyRelease>", self.autofill_first_level_spt)

        self.first_level_spt_listbox = tk.Listbox(self.first_level_filter_frame, height=3)
        self.first_level_spt_listbox.grid(row=1, column=3, padx=5, pady=0, sticky="ew")
        self.first_level_spt_listbox.bind('<<ListboxSelect>>', self.select_first_level_spt)
        self.first_level_spt_listbox.grid_remove() # Sembunyikan secara default

        self.display_first_level_button = tk.Button(self.first_level_filter_frame, text="Tampilkan Anak Langsung", command=self.display_first_level_bom_hierarchy)
        self.display_first_level_button.grid(row=0, column=4, padx=5, pady=5, sticky="w")

        self.restore_full_hierarchy_button = tk.Button(self.first_level_filter_frame, text="Kembali ke Tampilan Penuh", command=self.restore_full_hierarchy)
        self.restore_full_hierarchy_button.grid(row=0, column=5, padx=5, pady=5, sticky="w")

        # Bagian Input Nama Line
        # tk.Label(self.main_frame, text="Nama Line:").grid(row=6, column=0, padx=5, pady=5, sticky="w")
        # self.line_name_routing_entry = tk.Entry(self.main_frame, width=30)
        # self.line_name_routing_entry.grid(row=6, column=1, padx=5, pady=5, sticky="ew")
        # self.line_name_routing_entry.bind("<KeyRelease>", self.autofill_line_name)

        # self.line_name_listbox = tk.Listbox(self.main_frame, height=5)
        # self.line_name_listbox.grid(row=7, column=1, padx=5, pady=5, sticky="ew")
        # self.line_name_listbox.bind('<<ListboxSelect>>', self.select_line_name_from_list)
        # self.line_name_listbox.grid_remove() # Sembunyikan secara default

        # Tombol Simpan Routing
        # self.save_routing_button = tk.Button(self.main_frame, text="Simpan Routing", command=self.save_routing)
        # self.save_routing_button.grid(row=8, column=0, columnspan=3, pady=10) # Sesuaikan baris

        # Area tampilan hirarki BOM di Data Routing (dipindahkan ke atas)
        self.bom_hierarchy_frame = tk.LabelFrame(self.main_frame, text="Hirarki BOM Part Terpilih")
        self.bom_hierarchy_frame.grid(row=9, column=0, columnspan=3, padx=10, pady=5, sticky="nsew") # Pindahkan ke row 9

        # Frame untuk tombol-tombol
        self.bom_buttons_frame = tk.Frame(self.bom_hierarchy_frame)
        self.bom_buttons_frame.pack(fill="x", padx=5, pady=5)

        # Tombol Export ke Excel
        self.export_excel_button = tk.Button(self.bom_buttons_frame, text="Export ke Excel", command=self.export_bom_to_excel)
        self.export_excel_button.pack(side="left", padx=5)

        # Tombol Copy ke Clipboard
        self.copy_clipboard_button = tk.Button(self.bom_buttons_frame, text="Copy ke Clipboard", command=self.copy_bom_to_clipboard)
        self.copy_clipboard_button.pack(side="left", padx=5)

        # Frame untuk filter
        self.filter_frame = tk.Frame(self.bom_hierarchy_frame)
        self.filter_frame.pack(fill="x", padx=5, pady=5)
        
        # Label dan entry untuk filter
        tk.Label(self.filter_frame, text="Filter:").pack(side="left", padx=(0, 5))
        
        # Filter untuk setiap kolom
        self.filter_vars = {}
        self.filter_entries = {}
        
        filter_columns = ["Level", "Part No", "Description", "MRP Type", "MRP Controller", "SPT", "Line"]
        for i, col in enumerate(filter_columns):
            frame = tk.Frame(self.filter_frame)
            frame.pack(side="left", padx=2)
            
            tk.Label(frame, text=f"{col}:").pack()
            var = tk.StringVar()
            entry = tk.Entry(frame, textvariable=var, width=15)
            entry.pack()
            
            # Bind event untuk filter real-time
            var.trace('w', lambda *args, col=col: self.apply_filter(col))
            
            self.filter_vars[col] = var
            self.filter_entries[col] = entry
        
        # Tombol Clear Filter
        self.clear_filter_button = tk.Button(self.filter_frame, text="Clear Filter", command=self.clear_filters)
        self.clear_filter_button.pack(side="right", padx=5)

        # Frame untuk Treeview
        self.tree_frame = tk.Frame(self.bom_hierarchy_frame)
        self.tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Treeview untuk menampilkan data BOM dalam format tabel
        self.bom_tree = ttk.Treeview(self.tree_frame, columns=("Level", "Part No", "Description", "MRP Type", "MRP Controller", "SPT", "Line"), show="headings", height=15)
        
        # Konfigurasi kolom
        self.bom_tree.heading("Level", text="Level No")
        self.bom_tree.heading("Part No", text="Part No")
        self.bom_tree.heading("Description", text="Description")
        self.bom_tree.heading("MRP Type", text="MRP Type")
        self.bom_tree.heading("MRP Controller", text="MRP Controller")
        self.bom_tree.heading("SPT", text="SPT") # Tambah kolom SPT
        self.bom_tree.heading("Line", text="Line")
        
        # Atur lebar kolom dan alignment
        self.bom_tree.column("Level", width=60, anchor="w")  # Rata kiri
        self.bom_tree.column("Part No", width=100, anchor="center")  # Rata tengah
        self.bom_tree.column("Description", width=180, anchor="w")
        self.bom_tree.column("MRP Type", width=80, anchor="center")
        self.bom_tree.column("MRP Controller", width=100, anchor="center")
        self.bom_tree.column("SPT", width=80, anchor="center") # Atur lebar kolom SPT
        self.bom_tree.column("Line", width=100, anchor="center")
        
        # Scrollbar untuk Treeview
        self.tree_scrollbar = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.bom_tree.yview)
        self.bom_tree.configure(yscrollcommand=self.tree_scrollbar.set)
        
        # Pack Treeview dan Scrollbar
        self.bom_tree.pack(side="left", fill="both", expand=True)
        self.tree_scrollbar.pack(side="right", fill="y")

        # Bind double-click event
        self.bom_tree.bind("<Double-1>", self.on_tree_double_click)

        # Tambahkan ScrolledText sebagai backup (tersembunyi)
        self.bom_hierarchy_text = scrolledtext.ScrolledText(self.bom_hierarchy_frame, wrap=tk.WORD, width=60, height=15)
        self.bom_hierarchy_text.pack_forget()  # Sembunyikan

        # Frame untuk daftar routing yang tersimpan
        self.saved_routing_frame = tk.LabelFrame(self.main_frame, text="Daftar Routing Tersimpan")
        self.saved_routing_frame.grid(row=10, column=0, columnspan=3, padx=10, pady=10, sticky="nsew") # Pindahkan ke row 10

        self.routing_tree = ttk.Treeview(self.saved_routing_frame, columns=("Part No", "Line"), show="headings", height=6)
        self.routing_tree.heading("Part No", text="Part No")
        self.routing_tree.heading("Line", text="Line")
        self.routing_tree.column("Part No", width=200)
        self.routing_tree.column("Line", width=200)
        self.routing_tree.pack(side="left", fill="both", expand=True, padx=5, pady=5)

        self.delete_routing_button = tk.Button(self.saved_routing_frame, text="Hapus Routing Terpilih", command=self.delete_selected_routing)
        self.delete_routing_button.pack(side="left", padx=5, pady=5)

        self.load_routings()
        self.refresh_routings_list()

    def on_tree_double_click(self, event):
        region = self.bom_tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        column = self.bom_tree.identify_column(event.x)
        if column == "#6":  # Kolom "Line"
            item_id = self.bom_tree.identify_row(event.y)
            self.create_line_editor(item_id, column)

    def create_line_editor(self, item_id, column):
        # Dapatkan daftar line dari data_lines_manager
        line_names = [line['name'] for line in self.data_lines_manager.data_lines]
        if not line_names:
            messagebox.showinfo("Info", "Tidak ada data line yang tersedia. Harap tambahkan di menu Data Line.")
            return

        # Buat Combobox
        editor = ttk.Combobox(self.tree_frame, values=line_names, state="readonly")
        editor.place(relx=0.5, rely=0.5, anchor="center") # Atur posisi sementara

        # Dapatkan bounding box dari sel
        x, y, width, height = self.bom_tree.bbox(item_id, column)
        editor.place(x=x, y=y, width=width, height=height)

        # Bind event
        editor.bind("<<ComboboxSelected>>", lambda e: self.on_line_selected(e, item_id, editor))
        editor.focus_set()

    def on_line_selected(self, event, item_id, editor):
        new_value = editor.get()
        part_no = self.bom_tree.item(item_id, "values")[1]
        self.bom_tree.set(item_id, "Line", new_value)
        self.routings[part_no] = new_value
        self.save_routings()
        self.refresh_routings_list()
        editor.destroy()

    def load_bom_file(self):
        bom_directory = "N:\\Download\\BOM" # Lokasi file BOM yang dikoreksi
        latest_bom_file = None

        try:
            # Menggunakan glob_file_search untuk menemukan file .txt terbaru
            # Menggunakan glob_pattern "*.txt" dan target_directory
            # Asumsikan glob_file_search mengembalikan daftar file yang diurutkan berdasarkan waktu modifikasi (terbaru lebih dulu)
            # Atau kita perlu mengurutkannya secara manual jika tidak diurutkan.
            # Karena glob_file_search tidak menjamin urutan, kita akan memuat semua file yang cocok dan mencari yang terbaru.

            # Mencari semua file .txt di direktori
            import os
            import glob

            list_of_files = glob.glob(os.path.join(bom_directory, '*.txt'))
            if not list_of_files:
                messagebox.showwarning("Peringatan", f"Tidak ada file .txt ditemukan di {bom_directory}.")
                self.bom_hierarchy_text.insert(tk.END, "Tidak ada file .txt ditemukan untuk BOM.\n")
                return
            
            latest_file = max(list_of_files, key=os.path.getmtime)
            file_selected = latest_file

        except Exception as e:
            messagebox.showerror("Error", f"Terjadi kesalahan saat mencari file BOM terbaru:\n{e}")
            self.bom_hierarchy_text.insert(tk.END, f"Error mencari file BOM: {e}\n")
            return

        if file_selected:
            self.bom_file_path = file_selected
            self.bom_hierarchy_text.delete(1.0, tk.END)
            self.bom_hierarchy_text.insert(tk.END, f"Memuat data BOM dari {self.bom_file_path}...\n")
            print(f"[DEBUG] Mencoba memuat file BOM dari: {self.bom_file_path}") # Debugging
            try:
                self.bom_data = parse_bom_file(self.bom_file_path)
                if self.bom_data:
                    self.bom_hierarchy_text.insert(tk.END, "Data BOM berhasil dimuat.\n")
                    messagebox.showinfo("Sukses", "File BOM terbaru berhasil dimuat secara otomatis.")
                    print(f"[DEBUG] Jumlah entri BOM setelah dimuat: {len(self.bom_data)}") # Debugging
                else:
                    self.bom_hierarchy_text.insert(tk.END, "Gagal memuat data BOM atau file kosong.\n")
                    messagebox.showwarning("Peringatan", "Gagal memuat data BOM atau file kosong.")
                    print("[DEBUG] Data BOM kosong setelah parsing.") # Debugging
            except Exception as e:
                self.bom_hierarchy_text.insert(tk.END, f"Terjadi kesalahan saat memparsing file BOM: {e}\n")
                messagebox.showerror("Error", f"Terjadi kesalahan saat memparsing file BOM:\n{e}")
                print(f"[DEBUG] Error parsing BOM file: {e}") # Debugging

    def load_routings(self):
        try:
            with open("routing_data.json", "r") as f:
                self.routings = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self.routings = {}

    def save_routings(self):
        with open("routing_data.json", "w") as f:
            json.dump(self.routings, f, indent=4)

    def refresh_routings_list(self):
        """Segarkan daftar routing yang tersimpan di Treeview."""
        # Hapus item lama
        for item_id in self.routing_tree.get_children():
            self.routing_tree.delete(item_id)

        # Masukkan data dari self.routings (dict: part_no -> line)
        try:
            for part_no, line_name in sorted(self.routings.items()):
                self.routing_tree.insert("", "end", values=(part_no, line_name))
        except AttributeError:
            # Jika self.routings belum ada (seharusnya tidak terjadi setelah __init__), abaikan
            pass

    def delete_selected_routing(self):
        """Hapus routing terpilih dari penyimpanan dan refresh tampilan."""
        selected_items = self.routing_tree.selection()
        if not selected_items:
            messagebox.showwarning("Peringatan", "Pilih routing yang ingin dihapus.")
            return

        item_id = selected_items[0]
        values = self.routing_tree.item(item_id, "values")
        if not values:
            return
        part_no = values[0]

        if part_no in self.routings:
            del self.routings[part_no]
            self.save_routings()
            self.refresh_routings_list()
            messagebox.showinfo("Sukses", f"Routing untuk '{part_no}' telah dihapus.")
        else:
            messagebox.showwarning("Peringatan", f"Routing untuk '{part_no}' tidak ditemukan.")

    def autofill_mrp_controller(self, event=None):
        search_term = self.mrp_controller_entry.get().strip().upper()
        matching_mrpc = []
        for material_info in self.material_data.values():
            mrpc_val = material_info.get('MRPC', '')
            if search_term and mrpc_val.startswith(search_term) and mrpc_val not in matching_mrpc:
                matching_mrpc.append(mrpc_val)
        
        self.mrp_controller_listbox.delete(0, tk.END)
        if matching_mrpc:
            for mrpc in sorted(matching_mrpc):
                self.mrp_controller_listbox.insert(tk.END, mrpc)
            self.mrp_controller_listbox.grid()
        else:
            self.mrp_controller_listbox.grid_remove()

    def select_mrp_controller(self, event):
        if self.mrp_controller_listbox.curselection():
            selected_mrp = self.mrp_controller_listbox.get(self.mrp_controller_listbox.curselection())
            self.mrp_controller_entry.delete(0, tk.END)
            self.mrp_controller_entry.insert(0, selected_mrp)
            self.mrp_controller_listbox.grid_remove()
            self.display_parts_for_mrp_controller() # Langsung tampilkan parts setelah memilih MRPC

    def display_parts_for_mrp_controller(self):
        mrp_controller = self.mrp_controller_entry.get().strip()
        if not mrp_controller:
            messagebox.showwarning("Peringatan", "Silakan masukkan MRP Controller.")
            return

        # Hapus data lama
        for item in self.parts_tree.get_children():
            self.parts_tree.delete(item)

        related_parts = []
        for material, info in self.material_data.items():
            if info.get('MRPC') == mrp_controller:
                related_parts.append((material, info.get('Material Description', ''), info.get('SPT', '')))
        
        if related_parts:
            for part, description, spt in sorted(related_parts):
                self.parts_tree.insert("", "end", values=(part, description, spt))
        else:
            messagebox.showinfo("Info", f"Tidak ada parts yang ditemukan untuk MRP Controller: {mrp_controller}")

    def select_part_from_list(self, event):
        selected_item = self.parts_tree.selection()
        if selected_item:
            selected_part = self.parts_tree.item(selected_item, "values")[0]
            self.selected_part_entry.delete(0, tk.END)
            self.selected_part_entry.insert(0, selected_part)
            # Tampilkan MRP Type di kotak baru
            material_info = self.material_data.get(selected_part, {})
            mrp_type = material_info.get('MRP Type', '')
            self.selected_part_mrp_type_entry.config(state="normal")
            self.selected_part_mrp_type_entry.delete(0, tk.END)
            self.selected_part_mrp_type_entry.insert(0, mrp_type)
            self.selected_part_mrp_type_entry.config(state="readonly")

    def display_bom_hierarchy_for_part(self, part_induk):
        # Bersihkan Treeview
        for item in self.bom_tree.get_children():
            self.bom_tree.delete(item)
            
        print(f"[DEBUG] Mencoba menampilkan hirarki untuk Part: {part_induk}") # Debugging

        # Auto-load BOM jika belum dimuat
        if not self.bom_data:
            print("[DEBUG] BOM data kosong, mencoba memuat otomatis...")
            self.load_bom_file()
            if not self.bom_data:
                messagebox.showwarning("Peringatan", "Silakan muat file BOM terlebih dahulu.")
                print("[DEBUG] BOM data tetap kosong setelah percobaan muat otomatis.")
                return

        if not part_induk:
            messagebox.showwarning("Peringatan", "Silakan pilih Part terlebih dahulu.")
            print("[DEBUG] Part induk kosong, tidak bisa menampilkan hirarki.") # Debugging
            return
        
        part_found_as_parent = part_induk in self.bom_data
        part_found_as_child = any(part_induk in children for children in self.bom_data.values())
        print(f"[DEBUG] Part ditemukan sebagai parent: {part_found_as_parent}, Part ditemukan sebagai child: {part_found_as_child}") # Debugging

        if part_found_as_parent or part_found_as_child:
            # Kumpulkan semua data untuk filtering
            all_data = []
            
            # Tambahkan part induk
            material_info = self.material_data.get(part_induk, {})
            level = "1"  # Level 1 tanpa titik
            description = material_info.get('Material Description', '')
            mrp_type = material_info.get('MRP Type', '')
            mrpc = material_info.get('MRPC', '')
            spt = material_info.get('SPT', '') # Ambil nilai SPT
            line = self.routings.get(part_induk, "") # Ambil line dari data routing
            
            all_data.append((level, part_induk, description, mrp_type, mrpc, spt, line))
            
            # Tambahkan anak-anak jika ada
            if part_induk in self.bom_data:
                self.collect_child_data(part_induk, 2, all_data)
            
            # Simpan data asli untuk filtering
            self.original_bom_data = all_data
            
            # Tampilkan semua data ke Treeview
            for item_data in all_data:
                self.bom_tree.insert("", "end", values=item_data)

            # Pastikan filter kosong tidak menyembunyikan data
            self.clear_filters()
                
        else:
            # Tetap tampilkan informasi part minimal jika ada di material_data
            material_info = self.material_data.get(part_induk, {})
            if material_info:
                level = "1"
                description = material_info.get('Material Description', '')
                mrp_type = material_info.get('MRP Type', '')
                mrpc = material_info.get('MRPC', '')
                spt = material_info.get('SPT', '') # Ambil nilai SPT
                line = self.routings.get(part_induk, "")
                self.original_bom_data = [(level, part_induk, description, mrp_type, mrpc, spt, line)]
                self.bom_tree.insert("", "end", values=self.original_bom_data[0])
                self.clear_filters()
                messagebox.showinfo("Info", f"Part '{part_induk}' tidak memiliki anak dalam BOM. Menampilkan informasi part saja.")
                print(f"[DEBUG] Part {part_induk} tidak memiliki anak, menampilkan baris tunggal.")
            else:
                messagebox.showwarning("Peringatan", f"Part '{part_induk}' tidak ditemukan dalam data BOM maupun material.")
                print(f"[DEBUG] Part {part_induk} tidak ditemukan dalam BOM maupun material_data.")

    def collect_child_data(self, part, level, data_list):
        """Mengumpulkan data child secara rekursif untuk filtering"""
        for child in self.bom_data[part]:
            material_info = self.material_data.get(child, {})
            description = material_info.get('Material Description', '')
            mrp_type = material_info.get('MRP Type', '')
            mrpc = material_info.get('MRPC', '')
            spt = material_info.get('SPT', '') # Ambil nilai SPT
            line = self.routings.get(child, "") # Ambil line dari data routing
            
            # Format level dengan titik-titik sesuai hierarki
            # Level 1: "1", Level 2: "..2", Level 3: "...3", dst
            level_dots = "." * (level - 1) if level > 1 else ""
            formatted_level = f"{level_dots}{level}"
            
            data_list.append((formatted_level, child, description, mrp_type, mrpc, spt, line))
            
            # Tambahkan anak-anak jika ada
            if child in self.bom_data:
                self.collect_child_data(child, level + 1, data_list)

    def add_child_to_tree(self, part, level):
        """Menambahkan child part ke Treeview dengan level yang sesuai"""
        material_info = self.material_data.get(part, {})
        description = material_info.get('Material Description', '')
        mrp_type = material_info.get('MRP Type', '')
        mrpc = material_info.get('MRPC', '')
        
        # Format level dengan titik-titik sesuai hierarki
        # Level 1: "1", Level 2: "..2", Level 3: "...3", dst
        level_dots = "." * (level - 1) if level > 1 else ""
        formatted_level = f"{level_dots}{level}"
        
        self.bom_tree.insert("", "end", values=(formatted_level, part, description, mrp_type, mrpc))
        
        # Tambahkan anak-anak jika ada
        if part in self.bom_data:
            for child in self.bom_data[part]:
                self.add_child_to_tree(child, level + 1)

    def export_bom_to_excel(self):
        """Export data BOM dari Treeview ke file Excel"""
        if not self.bom_tree.get_children():
            messagebox.showwarning("Peringatan", "Tidak ada data BOM untuk di-export.")
            return
            
        file_path = filedialog.asksaveasfilename(
            title="Simpan File Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                # Buat workbook baru
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "BOM Hierarchy"
                
                # Header
                headers = ["Level No", "Part No", "Description", "MRP Type", "MRP Controller", "SPT", "Line"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                    ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
                
                # Data dari Treeview
                for row, item in enumerate(self.bom_tree.get_children(), 2):
                    values = self.bom_tree.item(item)['values']
                    for col, value in enumerate(values, 1):
                        ws.cell(row=row, column=col, value=value)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(file_path)
                messagebox.showinfo("Sukses", f"Data BOM berhasil di-export ke {file_path}")
                
            except Exception as e:
                messagebox.showerror("Error", f"Terjadi kesalahan saat export ke Excel:\n{e}")

    def copy_bom_to_clipboard(self):
        """Copy data BOM dari Treeview ke clipboard"""
        if not self.bom_tree.get_children():
            messagebox.showwarning("Peringatan", "Tidak ada data BOM untuk di-copy.")
            return
            
        try:
            # Header
            clipboard_text = "Level No\tPart No\tDescription\tMRP Type\tMRP Controller\tSPT\tLine\n"

            # Data dari Treeview
            for item in self.bom_tree.get_children():
                values = self.bom_tree.item(item)['values']
                clipboard_text += "\t".join(str(value) for value in values) + "\n"

            pyperclip.copy(clipboard_text)
            messagebox.showinfo("Sukses", "Data BOM berhasil di-copy ke clipboard.\nAnda dapat paste langsung ke Excel.")

        except Exception as e:
            messagebox.showerror("Error", f"Terjadi kesalahan saat copy ke clipboard:\n{e}")

    def autofill_selected_part(self, event=None):
        search_term = self.selected_part_entry.get().strip().upper()
        matching_parts = []
        for part, material_info in self.material_data.items():
            if search_term and part.startswith(search_term) and part not in matching_parts:
                matching_parts.append(part)
        
        self.selected_part_listbox.delete(0, tk.END)
        if matching_parts:
            for part in sorted(matching_parts):
                self.selected_part_listbox.insert(tk.END, part)
            self.selected_part_listbox.grid()
        else:
            self.selected_part_listbox.grid_remove()

    def select_selected_part(self, event):
        if self.selected_part_listbox.curselection():
            selected_part = self.selected_part_listbox.get(self.selected_part_listbox.curselection())
            self.selected_part_entry.delete(0, tk.END)
            self.selected_part_entry.insert(0, selected_part)
            
            material_info = self.material_data.get(selected_part, {})
            mrp_type = material_info.get('MRP Type', '')
            self.selected_part_mrp_type_entry.config(state="normal")
            self.selected_part_mrp_type_entry.delete(0, tk.END)
            self.selected_part_mrp_type_entry.insert(0, mrp_type)
            self.selected_part_mrp_type_entry.config(state="readonly")
            
            self.selected_part_listbox.grid_remove()

    def autofill_first_level_mrpc(self, event=None):
        search_term = self.first_level_mrpc_filter_entry.get().strip().upper()
        matching_mrpc = []
        for material_info in self.material_data.values():
            mrpc_val = material_info.get('MRPC', '')
            if search_term and mrpc_val.startswith(search_term) and mrpc_val not in matching_mrpc:
                matching_mrpc.append(mrpc_val)
        
        self.first_level_mrpc_listbox.delete(0, tk.END)
        if matching_mrpc:
            for mrpc in sorted(matching_mrpc):
                self.first_level_mrpc_listbox.insert(tk.END, mrpc)
            self.first_level_mrpc_listbox.grid()
        else:
            self.first_level_mrpc_listbox.grid_remove()

    def select_first_level_mrpc(self, event):
        if self.first_level_mrpc_listbox.curselection():
            selected_mrpc = self.first_level_mrpc_listbox.get(self.first_level_mrpc_listbox.curselection())
            self.first_level_mrpc_filter_entry.delete(0, tk.END)
            self.first_level_mrpc_filter_entry.insert(0, selected_mrpc)
            self.first_level_mrpc_listbox.grid_remove()

    def autofill_first_level_spt(self, event=None):
        search_term = self.first_level_spt_filter_entry.get().strip().upper()
        matching_spt = []
        for material_info in self.material_data.values():
            spt_val = material_info.get('SPT', '')
            if search_term and spt_val.startswith(search_term) and spt_val not in matching_spt:
                matching_spt.append(spt_val)
        
        self.first_level_spt_listbox.delete(0, tk.END)
        if matching_spt:
            for spt in sorted(matching_spt):
                self.first_level_spt_listbox.insert(tk.END, spt)
            self.first_level_spt_listbox.grid()
        else:
            self.first_level_spt_listbox.grid_remove()

    def select_first_level_spt(self, event):
        if self.first_level_spt_listbox.curselection():
            selected_spt = self.first_level_spt_listbox.get(self.first_level_spt_listbox.curselection())
            self.first_level_spt_filter_entry.delete(0, tk.END)
            self.first_level_spt_filter_entry.insert(0, selected_spt)
            self.first_level_spt_listbox.grid_remove()

    def autofill_line_name(self, event=None):
        search_term = self.line_name_routing_entry.get().strip().lower()
        matching_lines = []
        for num, name in self.data_lines_manager.data_lines:
            if search_term and name.lower().startswith(search_term) and name not in matching_lines:
                matching_lines.append(name)
        
        self.line_name_listbox.delete(0, tk.END)
        if matching_lines:
            for line_name in sorted(matching_lines):
                self.line_name_listbox.insert(tk.END, line_name)
            self.line_name_listbox.grid()
        else:
            self.line_name_listbox.grid_remove()

    def select_line_name_from_list(self, event):
        if self.line_name_listbox.curselection():
            selected_line = self.line_name_listbox.get(self.line_name_listbox.curselection())
            self.line_name_routing_entry.delete(0, tk.END)
            self.line_name_routing_entry.insert(0, selected_line)
            self.line_name_listbox.grid_remove()

    def save_routing(self):
        mrp_controller = self.mrp_controller_entry.get().strip()
        selected_part = self.selected_part_entry.get().strip()
        line_name = self.line_name_routing_entry.get().strip()

        if not mrp_controller or not selected_part or not line_name:
            messagebox.showwarning("Peringatan", "Semua kolom harus diisi.")
            return
        
        # Cek apakah line_name ada di data_lines_manager
        line_exists = False
        for num, name in self.data_lines_manager.data_lines:
            if name == line_name:
                line_exists = True
                break
        
        if not line_exists:
            messagebox.showwarning("Peringatan", f"Nama Line \'{line_name}\' tidak ditemukan di Data Line yang sudah ada. Harap tambahkan terlebih dahulu.")
            return

        # Tambahkan ke daftar routing
        routing_entry = {
            "MRP Controller": mrp_controller,
            "Part": selected_part,
            "Line Name": line_name
        }
        self.routings.append(routing_entry)
        self.save_routings()
        self.refresh_routings_list()
        messagebox.showinfo("Sukses", "Data Routing berhasil disimpan.")

    def apply_filter(self, column=None):
        """Terapkan filter ke Treeview"""
        # Hapus semua item yang ada
        for item in self.bom_tree.get_children():
            self.bom_tree.delete(item)
        
        # Jika tidak ada data asli, return
        if not hasattr(self, 'original_bom_data') or not self.original_bom_data:
            return
        
        # Ambil nilai filter dari semua kolom
        filters = {}
        for col in ["Level", "Part No", "Description", "MRP Type", "MRP Controller", "SPT", "Line"]:
            filters[col] = self.filter_vars[col].get().strip().lower()
        
        # Filter data berdasarkan semua kolom
        filtered_data = []
        for item_data in self.original_bom_data:
            # item_data adalah tuple: (level, part, description, mrp_type, mrpc, spt, line)
            level, part, description, mrp_type, mrpc, spt, line = item_data
            
            # Cek apakah item memenuhi semua filter
            matches = True
            
            # Filter khusus untuk MRPC dengan dukungan wildcard
            mrpc_filter_val = filters["MRP Controller"]
            if mrpc_filter_val:
                mrpc_normalized = str(mrpc).strip().upper() if mrpc else ""
                mrpc_filter_upper = mrpc_filter_val.upper()
                
                if mrpc_filter_upper == "*":
                    # Wildcard - semua MRPC cocok
                    pass
                elif "*" in mrpc_filter_upper:
                    # Pattern matching dengan wildcard
                    import fnmatch
                    if not fnmatch.fnmatch(mrpc_normalized, mrpc_filter_upper):
                        matches = False
                elif mrpc_filter_val.startswith("!"):
                    # Exclude filter
                    if mrpc_filter_val[1:].upper() in mrpc_normalized:
                        matches = False
                else:
                    # Exact match
                    if mrpc_filter_upper not in mrpc_normalized:
                        matches = False
            
            # Filter khusus untuk SPT dengan logika baru
            spt_filter_val = filters["SPT"]
            if spt_filter_val:
                spt_normalized = str(spt).strip().upper() if spt else ""
                spt_filter_upper = spt_filter_val.upper()
                
                if spt_filter_upper == "*":
                    # Wildcard - semua SPT cocok
                    pass
                elif not spt_filter_upper or spt_filter_upper == "":
                    # Jika filter kosong, hanya SPT kosong/blank yang cocok
                    if spt_normalized:
                        matches = False
                elif spt_filter_upper in ["BLANK", "KOSONG"]:
                    # Hanya SPT kosong/blank yang cocok
                    if spt_normalized:
                        matches = False
                elif spt_filter_val.startswith("!"):
                    # Exclude filter
                    if spt_filter_val[1:].upper() in spt_normalized:
                        matches = False
                else:
                    # Exact match
                    if spt_filter_upper not in spt_normalized:
                        matches = False
            
            # Filter untuk kolom lainnya (tidak termasuk MRPC dan SPT yang sudah ditangani khusus)
            for col_name, (val, filter_val) in zip(
                ["Level", "Part No", "Description", "MRP Type", "Line"],
                [
                    (level, filters["Level"]),
                    (part, filters["Part No"]),
                    (description, filters["Description"]),
                    (mrp_type, filters["MRP Type"]),
                    (line, filters["Line"])
                ]
            ):
                if filter_val:
                    if filter_val.startswith("!"):
                        if filter_val[1:] in str(val).lower():
                            matches = False
                            break
                    elif filter_val not in str(val).lower():
                        matches = False
                        break
            
            if matches:
                filtered_data.append(item_data)
        
        # Masukkan data yang sudah difilter ke Treeview
        for item_data in filtered_data:
            self.bom_tree.insert("", "end", values=item_data)

    def clear_filters(self):
        """Hapus semua filter dan tampilkan data asli"""
        # Clear semua entry filter
        for var in self.filter_vars.values():
            var.set("")
        
        # Tampilkan kembali data asli
        if hasattr(self, 'original_bom_data') and self.original_bom_data:
            # Hapus semua item yang ada
            for item in self.bom_tree.get_children():
                self.bom_tree.delete(item)
            
            # Masukkan kembali data asli
            for item_data in self.original_bom_data:
                self.bom_tree.insert("", "end", values=item_data)

    def muat_material_plant_data(self):
        """Memuat ulang data material plant dari file lokal atau dari MainMenuApp."""
        if self.main_app_instance:
            self.main_app_instance.reload_material_data()
            # Perbarui referensi material_data di jendela ini juga
            self.material_data = self.main_app_instance.material_data
            messagebox.showinfo("Info", "Proses pembaruan data material selesai.")
            # Setelah memuat ulang, jika ada tampilan yang bergantung pada material_data,
            # seperti daftar parts, mungkin perlu di-refresh.
            # self.display_parts_for_mrp_controller() # Ini akan refresh parts list jika MRP Controller sudah dipilih
        else:
            messagebox.showerror("Error", "Referensi ke aplikasi utama tidak ditemukan.")

    def display_first_level_bom_hierarchy(self):
        """Menampilkan anak terdekat dari part induk yang dipilih yang sesuai dengan filter MRPC dan SPT."""
        part_induk = self.selected_part_entry.get().strip()
        if not part_induk:
            messagebox.showwarning("Peringatan", "Silakan pilih Part Induk terlebih dahulu.")
            return

        # Pastikan BOM data sudah dimuat
        if not self.bom_data:
            messagebox.showwarning("Peringatan", "Silakan muat file BOM terlebih dahulu.")
            return

        if part_induk in self.bom_data:
            # Hapus semua item yang ada di treeview
            for item in self.bom_tree.get_children():
                self.bom_tree.delete(item)
            
            # Ambil nilai filter
            filter_mrpc = self.first_level_mrpc_filter_entry.get().strip().upper()
            filter_spt = self.first_level_spt_filter_entry.get().strip().upper()

            # List untuk menyimpan hasil anak terdekat yang sesuai filter
            filtered_nearest_children = []

            # Gunakan fungsi rekursif untuk mencari anak terdekat yang sesuai filter
            self._find_nearest_filtered_children(part_induk, filter_mrpc, filter_spt, filtered_nearest_children, current_level=1)

            # Simpan data asli untuk bisa dikembalikan nanti
            self.original_bom_data_backup = getattr(self, 'original_bom_data', [])
            
            # Simpan data anak terdekat sebagai data asli sementara
            self.original_bom_data = filtered_nearest_children

            # Tampilkan data yang sudah difilter dan diformat ke Treeview
            for item_data in filtered_nearest_children:
                self.bom_tree.insert("", "end", values=item_data)

            # Reset filter Treeview utama
            self.clear_filters()
            
            # Tampilkan pesan informasi
            if filtered_nearest_children:
                messagebox.showinfo("Info", f"Menampilkan {len(filtered_nearest_children)} anak terdekat untuk Part Induk '{part_induk}' (Terfilter MRPC dan SPT).")
            else:
                messagebox.showinfo("Info", f"Tidak ada anak terdekat untuk Part Induk '{part_induk}' yang sesuai dengan filter MRPC dan SPT.")
        else:
            messagebox.showwarning("Peringatan", f"Part Induk '{part_induk}' tidak memiliki anak dalam BOM.")

    def restore_full_hierarchy(self):
        """Mengembalikan tampilan hirarki penuh setelah menggunakan tampilan anak langsung."""
        if hasattr(self, 'original_bom_data_backup') and self.original_bom_data_backup:
            # Kembalikan data asli
            self.original_bom_data = self.original_bom_data_backup
            
            # Hapus semua item yang ada di treeview
            for item in self.bom_tree.get_children():
                self.bom_tree.delete(item)
            
            # Tampilkan kembali data asli
            for item_data in self.original_bom_data:
                self.bom_tree.insert("", "end", values=item_data)
            
            # Reset filter
            self.clear_filters()
            
            messagebox.showinfo("Info", "Tampilan hirarki penuh telah dikembalikan.")
        else:
            messagebox.showinfo("Info", "Tidak ada data hirarki penuh yang tersimpan.")

    def _find_nearest_filtered_children(self, current_part, filter_mrpc, filter_spt, found_children, current_level=1):
        """Fungsi rekursif untuk mencari anak terdekat yang cocok dengan filter MRPC dan SPT.
           Berhenti mencari di level yang lebih dalam jika sudah menemukan anak yang cocok di level saat ini.
        """
        children_of_current_part = self.bom_data.get(current_part, [])
        
        # List untuk menyimpan anak yang cocok di level saat ini
        matching_children_at_current_level = []

        for child in children_of_current_part:
            material_info = self.material_data.get(child, {})
            item_mrpc = material_info.get('MRPC', '').upper()
            item_spt = material_info.get('SPT', '').upper()

            # Filter MRPC dengan dukungan wildcard
            mrpc_match = False
            if not filter_mrpc or filter_mrpc == "*":
                # Jika filter kosong atau '*', tampilkan semua MRPC
                mrpc_match = True
            elif "*" in filter_mrpc:
                # Jika ada wildcard, gunakan pattern matching
                import fnmatch
                mrpc_match = fnmatch.fnmatch(item_mrpc, filter_mrpc)
            else:
                # Exact match
                mrpc_match = (filter_mrpc == item_mrpc)

            # Filter SPT dengan logika baru
            spt_match = False
            filter_spt_upper = filter_spt.upper()
            
            # Normalisasi nilai SPT item
            item_spt_normalized = item_spt.strip() if item_spt else ""
            
            if filter_spt_upper == "*": 
                # Jika filter '*', tampilkan semua SPT (wildcard)
                spt_match = True
            elif not filter_spt_upper or filter_spt_upper == "":
                # Jika filter kosong atau tidak terisi, tampilkan hanya SPT kosong/blank
                if not item_spt_normalized:
                    spt_match = True
            elif filter_spt_upper == "BLANK" or filter_spt_upper == "KOSONG":
                # Jika filter 'BLANK' atau 'KOSONG', tampilkan hanya SPT yang kosong/blank
                if not item_spt_normalized:
                    spt_match = True
            else:
                # Jika filter memiliki nilai spesifik, cocokkan dengan nilai SPT item
                if filter_spt_upper == item_spt_normalized.upper():
                    spt_match = True

            if mrpc_match and spt_match:
                # Jika cocok, tambahkan ke daftar anak yang cocok di level saat ini
                description = material_info.get('Material Description', '')
                mrp_type = material_info.get('MRP Type', '')
                line = self.routings.get(child, "")
                
                # Format level dengan titik-titik sesuai hierarki
                level_dots = "." * (current_level - 1) if current_level > 1 else ""
                formatted_level = f"{level_dots}{current_level}"
                
                matching_children_at_current_level.append((formatted_level, child, description, mrp_type, item_mrpc, item_spt, line))

        # Jika ada anak yang cocok di level saat ini, tambahkan ke hasil dan STOP mencari di level yang lebih dalam
        if matching_children_at_current_level:
            found_children.extend(matching_children_at_current_level)
        else:
            # Jika tidak ada anak yang cocok di level saat ini, lanjutkan mencari di level yang lebih dalam
            for child in children_of_current_part:
                self._find_nearest_filtered_children(child, filter_mrpc, filter_spt, found_children, current_level + 1)

    def _find_filtered_children_recursive(self, current_part, filter_mrpc, filter_spt, found_children, current_level=1):
        """Fungsi rekursif untuk mencari anak terdekat yang cocok dengan filter MRPC dan SPT.
           Jika parent_match_found True, kita mencari saudara-saudaranya di level yang sama.
        """
        children_of_current_part = self.bom_data.get(current_part, [])

        for child in children_of_current_part:
            material_info = self.material_data.get(child, {})
            item_mrpc = material_info.get('MRPC', '').upper()
            item_spt = material_info.get('SPT', '').upper()

            mrpc_match = (not filter_mrpc) or (filter_mrpc == item_mrpc)

            spt_match = False
            filter_spt_upper = filter_spt.upper()
            
            # Normalisasi nilai SPT item
            item_spt_normalized = item_spt.strip() if item_spt else ""
            
            if not filter_spt_upper or filter_spt_upper == "*": 
                # Jika filter kosong atau '*', tampilkan semua SPT (wildcard)
                spt_match = True
            elif filter_spt_upper == "BLANK" or filter_spt_upper == "KOSONG" or filter_spt_upper == "":
                # Jika filter 'BLANK', 'KOSONG', atau kosong, tampilkan hanya SPT yang kosong/blank
                if not item_spt_normalized:
                    spt_match = True
            else:
                # Jika filter memiliki nilai spesifik, cocokkan dengan nilai SPT item
                if filter_spt_upper == item_spt_normalized.upper():
                    spt_match = True

            if mrpc_match and spt_match:
                # Jika cocok, tambahkan ke daftar hasil dengan level hierarki yang sesuai
                description = material_info.get('Material Description', '')
                mrp_type = material_info.get('MRP Type', '')
                line = self.routings.get(child, "")
                
                level_dots = "." * (current_level - 1) if current_level > 1 else ""
                formatted_level = f"{level_dots}{current_level}"
                
                found_children.append((formatted_level, child, description, mrp_type, item_mrpc, item_spt, line))
                # JANGAN menelusuri lebih dalam karena kita hanya mencari anak terdekat yang cocok
            else:
                # Jika tidak cocok, lanjutkan pencarian di anak-anaknya dengan level yang dinaikkan
                self._find_filtered_children_recursive(child, filter_mrpc, filter_spt, found_children, current_level + 1)


if __name__ == "__main__":
    root = tk.Tk()
    app = MainMenuApp(root)
    root.mainloop()